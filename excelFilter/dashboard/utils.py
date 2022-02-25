import openpyxl
from pathlib import Path
from datetime import datetime

from . import config

DATASOURCE_PATH = Path(Path.home(), 'Documents', 'djangoprojects', 'excel-filter', 'datasource')

config_map = {
    'email-followup': config.FOLLOWUP_CONFIG,
    'email-file': config.FOR_FILE_CONFIG,
    'email-initial': config.INITIAL_CONFIG,
    'email-replies': config.EMAIL_REPLIES_CONFIG,
}

def get_file_data(request, menu_slug):
    file_data = {}
    file_config = config_map[menu_slug]
    hidden_cols = file_config['HIDDEN_COLUMNS']
    if menu_slug == 'email-followup':
        file_data['rows'] = _get_data('FOLLOW UP 2022.xlsx', file_config['DATA_START_ROW'], file_config['NON_EMPTY_COL'], hidden_cols)
        file_data['cols'] = file_config['TABLE_COLUMNS']
    elif menu_slug == 'email-file':
        file_data['rows'] = _get_data('FOR FILE 2022.xlsx', file_config['DATA_START_ROW'], file_config['NON_EMPTY_COL'], hidden_cols)
        file_data['cols'] = file_config['TABLE_COLUMNS']
    elif menu_slug == 'email-initial':
        file_data['rows'] = _get_data('INITIAL 2022.xlsx', file_config['DATA_START_ROW'], file_config['NON_EMPTY_COL'], hidden_cols)
        file_data['cols'] = file_config['TABLE_COLUMNS']
    elif menu_slug == 'email-replies':
        file_data['rows'] = _get_data('REPLY 2022.xlsx', file_config['DATA_START_ROW'], file_config['NON_EMPTY_COL'], hidden_cols)
        file_data['cols'] = file_config['TABLE_COLUMNS']
    post_params = get_post_params(request)
    if request.method == 'POST':
        if menu_slug == 'email-replies':
            file_data_from_daterange_1 = get_filedata_from_daterange(post_params[0][0], post_params[0][1], file_data,
                                                                     file_config['FILTER_FIELDS']['DATE_OF_EMAIL_COL'])

            file_data_from_daterange_2 = get_filedata_from_daterange(post_params[1][0], post_params[1][1], file_data_from_daterange_1,
                                                                     file_config['FILTER_FIELDS']['DATE_ENCODED_COL'])
            file_data = file_data_from_daterange_2
        else:
            file_data_from_daterange_1 = get_filedata_from_daterange(post_params[0][0], post_params[0][1], file_data,
                                                                     file_config['FILTER_FIELDS']['DATE_OF_EMAIL_COL'])

            file_data_from_daterange_2 = get_filedata_from_daterange(post_params[1][0], post_params[1][1],
                                                                     file_data_from_daterange_1,
                                                                     file_config['FILTER_FIELDS']['DATE_EVALUATED_COL'])
            file_data = file_data_from_daterange_2
    return file_data

# date_field: DATE_OF_EMAIL_COL or DATE_EVALUATED_COL
def get_filedata_from_daterange(from_date, to_date, file_data, date_field_col):
    file_data_rows = file_data['rows']
    rows_from_daterange = []
    if from_date and to_date:
        from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
        for row in file_data_rows:
            _date = row[date_field_col - 1].date()
            if from_date <= _date <= to_date:
                rows_from_daterange.append(row)
        file_data['rows'] = rows_from_daterange
    return file_data

def update_context(request, menu_slug, context):
    file_data = get_file_data(request, menu_slug)
    post_params = get_post_params(request)
    context['menu_slug'] = menu_slug
    context['rows'] = file_data['rows']
    context['cols'] = file_data['cols']
    context['data_controller_list'] = config.DATA_CONTROLLER_LIST
    context['nature_or_concern_list'] = config.NATURE_OR_CONCERN_LIST
    context['referred_to_agency_list'] = config.REFERRED_TO_AGENCY_LIST
    context['filter_columns'] = _get_filter_columns(menu_slug)
    context['post_params'] = post_params
    return context

def get_dashboard_overall_summary_today(request):
    dashboard_summary = {}
    for key, file_config in config_map.items():
        today_rows = []
        rows = get_file_data(request, key)['rows']
        if key == 'email-replies':
            date_col = file_config['FILTER_FIELDS']['DATE_ENCODED_COL'] - 1
        else:
            date_col = file_config['FILTER_FIELDS']['DATE_OF_EMAIL_COL'] - 1
        for row in rows:
            if datetime.now().date() == row[date_col].date():
                today_rows.append(row)
        dashboard_summary[key] = today_rows
    return dashboard_summary

def get_dashboard_controller_summary_today(request):

    # structure = {
    #       'AAA': [[controller_info], {
    #           'email-followup': [rows],
    #           'email-file': [rows],
    #           'email-initial': [rows],
    #           'email-replies': [rows],
    #       }]
    #       'BBB': [[controller_info], {
    #           'email-followup': [rows],
    #           'email-file': [rows],
    #           'email-initial': [rows],
    #           'email-replies': [rows],
    #       }]
    # }
    controller_overall_summary_today = {}
    dashboard_summary = get_dashboard_overall_summary_today(request)
    for controller in config.DATA_CONTROLLER_LIST:
        controller_initials = controller[0]
        controller_summary = {}
        for key, rows in dashboard_summary.items():
            controller_rows = []
            for row in rows:
                if key == 'email-replies':
                    controller_col = config_map[key]['FILTER_FIELDS']['EVALUATOR_OR_SENDER_ENCODER'] - 1
                else:
                    controller_col = config_map[key]['FILTER_FIELDS']['AO_E_SENDER_DATA_ENCODER'] - 1
                if controller_initials in _get_controller_list(row[controller_col]):
                    # print(key, controller_initials, row)
                    # print('--------------------------')
                    controller_rows.append(row)
            controller_summary[key] = controller_rows
        controller_overall_summary_today[controller_initials] = [controller, controller_summary]
    return controller_overall_summary_today

def get_controller_total_today(request):
    temp = get_dashboard_controller_summary_today(request)
    sums = []
    _sum = ''
    for key, values in temp.items():
        _sum = 0
        for key1, file_rows in values[1].items():
            _sum += len(file_rows)
        sums.append(_sum)
    return sums

def get_post_params(request):
    from_daterange_1 = request.POST.get('from-daterange-1')
    to_daterange_1 = request.POST.get('to-daterange-1')
    from_daterange_2 = request.POST.get('from-daterange-2')
    to_daterange_2 = request.POST.get('to-daterange-2')

    if not from_daterange_1 and to_daterange_1:
        from_daterange_1 = to_daterange_1
    elif from_daterange_1 and not to_daterange_1:
        to_daterange_1 = from_daterange_1

    if not from_daterange_2 and to_daterange_2:
        from_daterange_2 = to_daterange_2
    elif from_daterange_2 and not to_daterange_2:
        to_daterange_2 = from_daterange_2

    return [[from_daterange_1, to_daterange_1], [from_daterange_2, to_daterange_2]]

# PRIVATE METHODS

def _get_controller_list(row_controller_str):
    row_controller_list = row_controller_str.split('/')
    stripped_controller_list = []
    for controller in row_controller_list:
        stripped_controller_list.append(controller.strip())
    return stripped_controller_list

def _get_filter_columns(menu_slug):
    filter_columns = {}
    if menu_slug == 'email-followup':
        filter_columns = config.FOLLOWUP_CONFIG['FILTER_FIELDS']
    elif menu_slug == 'email-file':
        filter_columns = config.FOR_FILE_CONFIG['FILTER_FIELDS']
    elif menu_slug == 'email-initial':
        filter_columns = config.INITIAL_CONFIG['FILTER_FIELDS']
    elif menu_slug == 'email-replies':
        filter_columns = config.EMAIL_REPLIES_CONFIG['FILTER_FIELDS']
    return filter_columns

def _get_active_sheet(xlsx_file):
    xlsx_file_dir = DATASOURCE_PATH / xlsx_file
    wb_obj = openpyxl.load_workbook(xlsx_file_dir)
    sheet = wb_obj.active
    return sheet

def _get_data(file_name, data_row_start, non_empty_col, hidden_cols):
    sheet = _get_active_sheet(file_name)
    rows = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        # Check if Value of Date of Email is not empty
        _i = data_row_start - 1
        if i >= _i and sheet.cell(i + 1, non_empty_col).value is not None:
            row_as_list = list(row)
            for col in sorted(hidden_cols, reverse=True):
                del row_as_list[col]

            rows.append(tuple(row_as_list))
    return rows


